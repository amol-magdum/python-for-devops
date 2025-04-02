import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Load an existing workbook
workbook = openpyxl.load_workbook('example.xlsx')

# Select a worksheet by name
sheet = workbook['Sheet1']

# Read data from a specific cell
cell_value = sheet['A1'].value
print(f"Value in A1: {cell_value}")

cell2 = sheet['B2'].value
print(f"Value in B2: {cell2}")

# get the active sheet
print(workbook.active.title) 

# roe + column
value1 = sheet.cell(row=2, column=1).value
print(f"Value in row 2, column 1: {value1}")

# get max row and column count
print(f"Max row: {sheet.max_row}")
print(f"Max column: {sheet.max_column}")

# get columns name and row numbers
print(f"Column name: {get_column_letter(1)}")
print(f"Column number: {column_index_from_string('A')}")

print(f"row name: {get_(2)}")
# Close the workbook
workbook.close()