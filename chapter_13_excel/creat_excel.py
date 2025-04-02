import openpyxl

wb = openpyxl.Workbook()

print(wb.sheetnames)
print(wb.active.title)

sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet' # Change title.

print(wb.sheetnames)
print(wb.active.title)


#adding new sheet
wb.create_sheet('another food items')


#save the sheet to drive
wb.save('new.xlsx')

print(wb.sheetnames)